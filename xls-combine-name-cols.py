import openpyxl


input_file = "input.xlsx"
output_file = "output.xlsx"

# Required headers, case insensitive
xls_headers = [["first name",False,-1],["spouse/partner name",False,-1],["last name",False,-1]]


def validate_names(names):
    name_list = []
    for name in names:
        if name != None:
            name_strip = name.strip()
            if len(name_strip) > 0:
                name_list.append(name_strip)
    # If list contains 3+ items, insert "and"
    if len(name_list) >= 3:
        name_list.insert(1, "and")
    # Return combined names
    return " ".join(name_list)


print("Loading xls file...")

# Load input file
wb = openpyxl.load_workbook(input_file)
# Active workbook tab
ws = wb.active


print("Checking headers...")

# Find matching headers in xls file
for row_cells in ws.iter_rows(min_row=1, max_row=1):
    for pos, cell in enumerate(row_cells):
        for item in xls_headers:
            if cell.value.lower() == item[0].lower():
                item[1] = True
                item[2] = pos

# Check for any missing headers
if False in [i[1] for i in xls_headers]:
    for item in [i[0] for i in xls_headers if i[1] == False]:
        print(f"Missing header: {item}")
    raise SystemExit()


print("Combining names...")

# Find new column position
new_column_num = ws.max_column + 1
# Add header for new column
ws.cell(row=1, column=new_column_num, value="Combined Names")

# Loop through rows in workbook
for row_num, row in enumerate(ws[2:ws.max_row], start=2):

    # Get names from row, based on header index [first, spouse, last]
    combined_names = validate_names([row[xls_headers[0][2]].value, row[xls_headers[1][2]].value, row[xls_headers[2][2]].value])

    # Add new value to worksheet
    ws.cell(row=row_num, column=new_column_num, value=combined_names)


print("Saving file...")

# Save file
wb.save(output_file)
